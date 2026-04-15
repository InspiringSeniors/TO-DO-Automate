from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from uuid import UUID
from datetime import datetime, date, timezone
from zoneinfo import ZoneInfo

IST = ZoneInfo("Asia/Kolkata")
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.db import database, models
from app.schemas import task as task_schema
from app.api.dependencies import get_current_user, get_current_admin
from app.services import excel_parser

router = APIRouter()

@router.post("", response_model=task_schema.TaskOut)
def create_task(task: task_schema.TaskCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    new_task = models.Task(**task.model_dump(), user_id=current_user.id)
    db.add(new_task)
    db.commit()
    db.refresh(new_task)
    return new_task

@router.get("", response_model=list[task_schema.TaskOut])
def get_tasks(
    status: str = None, 
    priority: str = None,
    due_today: bool = False,
    overdue: bool = False,
    search: str = None,
    db: Session = Depends(database.get_db), 
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Task)
    
    # Standard employee sees own tasks. Admin sees all.
    if current_user.role != "admin":
        query = query.filter(models.Task.user_id == current_user.id)
        
    if status:
        query = query.filter(models.Task.status == status)
    if priority:
        query = query.filter(models.Task.priority == priority)
    if search:
        query = query.filter(models.Task.task_name.ilike(f"%{search}%"))
        
    now = datetime.now(timezone.utc)
    if overdue:
        query = query.filter(models.Task.due_datetime < now, models.Task.status != 'completed')
    if due_today:
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        today_end   = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59, microsecond=999999)
        query = query.filter(models.Task.due_datetime >= today_start, models.Task.due_datetime <= today_end)

    # Always sort by due date ascending (soonest first)
    query = query.order_by(models.Task.due_datetime.asc())
    return query.all()

@router.get("/export-excel")
def export_tasks_excel(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Download all tasks for the current user as a formatted .xlsx file."""
    query = db.query(models.Task)
    if current_user.role != "admin":
        query = query.filter(models.Task.user_id == current_user.id)
    tasks = query.order_by(models.Task.due_datetime.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Task Name", "Description", "Due Date", "Due Time", "Priority", "Status"]
    col_widths = [35, 40, 14, 13, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    priority_colors = {"high": "FEE2E2", "medium": "FEF3C7", "low": "DCFCE7"}
    status_colors   = {"completed": "DCFCE7", "pending": "FEF9C3", "in_progress": "DBEAFE"}

    for row_idx, task in enumerate(tasks, start=2):
        due_dt = task.due_datetime
        if due_dt:
            # Ensure timezone-aware, then convert to IST for display
            if due_dt.tzinfo is None:
                due_dt = due_dt.replace(tzinfo=timezone.utc)
            due_dt_ist    = due_dt.astimezone(IST)
            due_date_str  = due_dt_ist.strftime("%Y-%m-%d")
            due_time_str  = due_dt_ist.strftime("%I:%M %p")   # e.g. 02:30 PM
        else:
            due_date_str = due_time_str = ""

        values = [
            task.task_name,
            task.description or "",
            due_date_str,
            due_time_str,
            task.priority.capitalize() if task.priority else "",
            task.status.capitalize()   if task.status   else "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")

        p_color = priority_colors.get((task.priority or "").lower())
        s_color = status_colors.get((task.status or "").lower())
        if p_color:
            ws.cell(row=row_idx, column=5).fill = PatternFill(start_color=p_color, end_color=p_color, fill_type="solid")
        if s_color:
            ws.cell(row=row_idx, column=6).fill = PatternFill(start_color=s_color, end_color=s_color, fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tasks_{current_user.username}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{task_id}", response_model=task_schema.TaskOut)
def get_task(task_id: UUID, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task or (task.user_id != current_user.id and current_user.role != "admin"):
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@router.put("/{task_id}", response_model=task_schema.TaskOut)
def update_task(task_id: UUID, task_update: task_schema.TaskUpdate, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task or (task.user_id != current_user.id and current_user.role != "admin"):
        raise HTTPException(status_code=404, detail="Task not found")
        
    update_data = task_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(task, key, value)
        
    db.commit()
    db.refresh(task)
    return task

@router.delete("/{task_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_task(task_id: UUID, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task or (task.user_id != current_user.id and current_user.role != "admin"):
        raise HTTPException(status_code=404, detail="Task not found")
    
    db.delete(task)
    db.commit()
    return None

@router.patch("/{task_id}/complete", response_model=task_schema.TaskOut)
def complete_task(task_id: UUID, db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task or (task.user_id != current_user.id and current_user.role != "admin"):
        raise HTTPException(status_code=404, detail="Task not found")
        
    task.status = "completed"
    db.commit()
    db.refresh(task)
    return task

@router.post("/upload-excel")
async def upload_excel_tasks(file: UploadFile = File(...), db: Session = Depends(database.get_db), current_user: models.User = Depends(get_current_user)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    contents = await file.read()
    stats = excel_parser.process_tasks_excel(contents, current_user.id, db)
    return {"message": "Upload successful", "stats": stats}
