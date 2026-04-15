import smtplib
import logging
import os
import io
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

IST = timezone(timedelta(hours=5, minutes=30))


def _gmail():
    return os.getenv("GMAIL_EMAIL", ""), os.getenv("GMAIL_APP_PASSWORD", "")


def _to_ist(dt) -> datetime:
    """Convert any datetime to IST. Treats naive datetimes as UTC."""
    if dt is None:
        return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST)


def _fmt(dt) -> str:
    """Format a datetime in IST as '15 Apr 2026, 11:52 AM'."""
    if dt is None:
        return "—"
    return _to_ist(dt).strftime("%d %b %Y, %I:%M %p")


def _smtp_send(msg) -> bool:
    gmail, app_pw = _gmail()
    if not gmail or not app_pw:
        logger.warning("Gmail credentials not configured – skipping.")
        return False
    for attempt in range(3):
        try:
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as srv:
                srv.ehlo()
                srv.starttls()
                srv.login(gmail, app_pw)
                srv.sendmail(gmail, msg["To"], msg.as_string())
            return True
        except Exception as e:
            logger.error(f"SMTP attempt {attempt + 1} failed: {e}")
    return False


# ── HTML builders ──────────────────────────────────────────────────────────────

def _build_reminder_html(employee_name: str, task_name: str, due_datetime, priority: str) -> str:
    colors = {"high": "#ef4444", "medium": "#f59e0b", "low": "#22c55e"}
    color   = colors.get(priority.lower(), "#6366f1")
    due_str = _fmt(due_datetime)          # ← IST with AM/PM
    return f"""
    <html><head><style>
      body{{font-family:Arial,sans-serif;background:#f1f5f9;margin:0;padding:0;}}
      .wrap{{max-width:520px;margin:40px auto;background:#fff;border-radius:12px;
             box-shadow:0 4px 20px rgba(0,0,0,.08);overflow:hidden;}}
      .hdr{{background:#6366f1;padding:28px 32px;}}
      .hdr h1{{color:#fff;margin:0;font-size:20px;}}
      .body{{padding:28px 32px;}}
      .badge{{display:inline-block;padding:4px 12px;border-radius:999px;
              background:{color}22;color:{color};font-size:13px;font-weight:600;text-transform:capitalize;}}
      .box{{background:#f8fafc;border-left:4px solid {color};border-radius:8px;padding:16px 20px;margin:20px 0;}}
      .box p{{margin:0 0 6px;color:#475569;font-size:14px;}}
      .box h2{{margin:0 0 12px;color:#1e293b;font-size:18px;}}
      .foot{{padding:16px 32px;background:#f8fafc;color:#94a3b8;font-size:12px;}}
    </style></head>
    <body><div class="wrap">
      <div class="hdr"><h1>⏰ Task Reminder</h1></div>
      <div class="body">
        <p>Hi <strong>{employee_name}</strong>,</p>
        <p>You have a task due in the next <strong>30 minutes</strong>.</p>
        <div class="box">
          <p>Task</p><h2>{task_name}</h2>
          <p>📅 Due: <strong>{due_str}</strong></p>
          <p>Priority: <span class="badge">{priority}</span></p>
        </div>
        <p>Log in to the dashboard to mark it as done.</p>
      </div>
      <div class="foot">Automated reminder from TO-DO Automate — do not reply.</div>
    </div></body></html>"""


def _build_digest_html(employee_name: str, tasks: list, heading: str,
                       has_attachment: bool = False) -> str:
    colors = {"high": "#ef4444", "medium": "#f59e0b", "low": "#22c55e"}
    rows = ""
    for t in tasks:
        color   = colors.get((t.priority or "medium").lower(), "#6366f1")
        due_str = _fmt(t.due_datetime)    # ← IST with AM/PM
        rows += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #f1f5f9;color:#1e293b;font-size:14px;">{t.task_name}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f1f5f9;color:#64748b;font-size:13px;">{due_str}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f1f5f9;text-align:center;">
            <span style="background:{color}22;color:{color};padding:3px 10px;border-radius:999px;
                         font-size:12px;font-weight:600;text-transform:capitalize;">{t.priority or 'medium'}</span>
          </td>
        </tr>"""

    attachment_note = (
        "<p style='color:#475569;font-size:14px;'>"
        "📎 Your full pending task list is <strong>attached as an Excel file</strong> for easy reference.</p>"
    ) if has_attachment else ""

    return f"""
    <html><head></head>
    <body style="font-family:Arial,sans-serif;background:#f1f5f9;margin:0;padding:0;">
      <div style="max-width:560px;margin:40px auto;background:#fff;border-radius:12px;
                  box-shadow:0 4px 20px rgba(0,0,0,.08);overflow:hidden;">
        <div style="background:#6366f1;padding:28px 32px;">
          <h1 style="color:#fff;margin:0;font-size:20px;">{heading}</h1>
        </div>
        <div style="padding:28px 32px;">
          <p style="color:#334155;">Hi <strong>{employee_name}</strong>,</p>
          <p style="color:#475569;font-size:14px;">
            You have <strong>{len(tasks)}</strong> pending task{'s' if len(tasks) != 1 else ''}.
          </p>
          {attachment_note}
          <table style="width:100%;border-collapse:collapse;margin:16px 0;
                         border:1px solid #e2e8f0;border-radius:8px;overflow:hidden;">
            <thead><tr style="background:#f8fafc;">
              <th style="padding:10px 14px;text-align:left;font-size:12px;color:#64748b;text-transform:uppercase;">Task</th>
              <th style="padding:10px 14px;text-align:left;font-size:12px;color:#64748b;text-transform:uppercase;">Due (IST)</th>
              <th style="padding:10px 14px;text-align:center;font-size:12px;color:#64748b;text-transform:uppercase;">Priority</th>
            </tr></thead>
            <tbody>{rows}</tbody>
          </table>
          <p style="color:#475569;font-size:14px;">Log in to the dashboard to manage your tasks.</p>
        </div>
        <div style="padding:16px 32px;background:#f8fafc;color:#94a3b8;font-size:12px;">
          Automated digest from TO-DO Automate — do not reply.
        </div>
      </div>
    </body></html>"""


# ── Excel builder ──────────────────────────────────────────────────────────────

def _build_pending_excel(tasks: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Tasks"

    hdr_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers  = ["#", "Task Name", "Description", "Due Date (IST)", "Due Time (IST)", "Priority", "Status"]
    widths   = [5, 35, 38, 16, 15, 12, 12]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    p_colors = {"high": "FEE2E2", "medium": "FEF3C7", "low": "DCFCE7"}

    for idx, task in enumerate(tasks, start=1):
        ist_dt = _to_ist(task.due_datetime)
        row_vals = [
            idx,
            task.task_name,
            task.description or "",
            ist_dt.strftime("%Y-%m-%d") if ist_dt else "",
            ist_dt.strftime("%I:%M %p") if ist_dt else "",   # ← 12-hour IST
            task.priority.capitalize() if task.priority else "",
            task.status.capitalize()   if task.status   else "",
        ]
        r = idx + 1
        for col, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=col, value=val).alignment = Alignment(vertical="center")

        pc = p_colors.get((task.priority or "").lower())
        if pc:
            ws.cell(row=r, column=6).fill = PatternFill(start_color=pc, end_color=pc, fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public send functions ──────────────────────────────────────────────────────

def send_reminder_email(to_email: str, employee_name: str, task_name: str,
                        due_datetime, priority: str) -> bool:
    gmail, _ = _gmail()
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"⏰ Reminder: '{task_name}' is due in 30 minutes"
    msg["From"]    = gmail
    msg["To"]      = to_email
    msg.attach(MIMEText(_build_reminder_html(employee_name, task_name, due_datetime, priority), "html"))
    ok = _smtp_send(msg)
    if ok:
        logger.info(f"Reminder sent → {to_email} | task: {task_name}")
    return ok


def send_digest_email(to_email: str, employee_name: str, tasks: list) -> bool:
    """3 PM afternoon digest — HTML only."""
    gmail, _ = _gmail()
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"📋 {len(tasks)} overdue task{'s' if len(tasks) != 1 else ''} still pending — TO-DO Automate"
    msg["From"]    = gmail
    msg["To"]      = to_email
    msg.attach(MIMEText(_build_digest_html(employee_name, tasks, "📋 Overdue Task Digest"), "html"))
    ok = _smtp_send(msg)
    if ok:
        logger.info(f"Afternoon digest → {to_email} ({len(tasks)} tasks)")
    return ok


def send_morning_digest_with_excel(to_email: str, employee_name: str,
                                    tasks: list, username: str) -> bool:
    """9:30 AM morning digest — HTML body + Excel attachment."""
    gmail, _ = _gmail()
    today_str = datetime.now(IST).strftime("%d %b %Y")

    # Correct MIME structure for HTML + attachment:
    # multipart/mixed
    #   ├── multipart/alternative
    #   │     └── text/html
    #   └── application/xlsx  (attachment)
    outer = MIMEMultipart("mixed")
    outer["Subject"] = f"📅 Good Morning {employee_name.split()[0]}! Pending tasks — {today_str}"
    outer["From"]    = gmail
    outer["To"]      = to_email

    # HTML wrapped in alternative
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(
        _build_digest_html(employee_name, tasks,
                           "📅 Good Morning — Your Pending Tasks",
                           has_attachment=True),
        "html"
    ))
    outer.attach(alt)

    # Excel attachment
    excel_bytes = _build_pending_excel(tasks)
    part = MIMEBase("application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    fname = f"pending_tasks_{username}_{datetime.now(IST).strftime('%Y%m%d')}.xlsx"
    part.add_header("Content-Disposition", "attachment", filename=fname)
    part.add_header("Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    name=fname)
    outer.attach(part)

    ok = _smtp_send(outer)
    if ok:
        logger.info(f"Morning digest + Excel → {to_email} ({len(tasks)} tasks)")
    return ok
